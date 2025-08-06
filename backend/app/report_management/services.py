"""
Report Management Services

This module provides comprehensive report generation and data export
capabilities for the Arusha Catholic Seminary School Management System.
"""

import json
import os
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Union
from pathlib import Path

import pandas as pd
from sqlalchemy.orm import Session
from sqlalchemy import func
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill

from ..models import User, Student, Teacher, Class, Grade, Attendance
from .models import ReportTemplate, ReportLog, ReportSchedule, ReportType, ReportFormat, ReportStatus


class ReportService:
    """Service class for report generation and management"""
    
    def __init__(self, db: Session):
        self.db = db
        self.report_dir = Path("reports")
        self.report_dir.mkdir(exist_ok=True)
        self.styles = getSampleStyleSheet()
    
    async def create_template(self, template_data: Dict[str, Any], created_by: int) -> ReportTemplate:
        """Create a new report template"""
        template = ReportTemplate(
            name=template_data["name"],
            description=template_data.get("description"),
            report_type=template_data["report_type"],
            template_config=template_data["template_config"],
            query_config=template_data["query_config"],
            output_formats=template_data["output_formats"],
            parameters=template_data.get("parameters"),
            created_by=created_by
        )
        
        self.db.add(template)
        self.db.commit()
        self.db.refresh(template)
        return template
    
    async def get_templates(self, skip: int = 0, limit: int = 100) -> List[ReportTemplate]:
        """Get report templates"""
        return self.db.query(ReportTemplate).filter(
            ReportTemplate.is_active == True
        ).offset(skip).limit(limit).all()
    
    async def generate_report(self, template_id: Optional[int], report_data: Dict[str, Any], created_by: int) -> ReportLog:
        """Generate a report"""
        start_time = datetime.utcnow()
        
        report_log = ReportLog(
            template_id=template_id,
            report_name=report_data["report_name"],
            report_type=report_data["report_type"],
            output_format=report_data["output_format"],
            status=ReportStatus.PENDING,
            parameters=report_data.get("parameters"),
            created_by=created_by
        )
        
        self.db.add(report_log)
        self.db.commit()
        self.db.refresh(report_log)
        
        try:
            data = await self._get_report_data(report_data["report_type"], report_data.get("parameters", {}))
            file_path, file_size = await self._generate_output(
                data, report_data["output_format"], report_data["report_name"]
            )
            
            report_log.status = ReportStatus.COMPLETED
            report_log.file_path = str(file_path)
            report_log.file_size = file_size
            report_log.processing_time = (datetime.utcnow() - start_time).total_seconds()
            report_log.record_count = len(data) if isinstance(data, list) else 1
            report_log.generated_at = datetime.utcnow()
            
        except Exception as e:
            report_log.status = ReportStatus.FAILED
            report_log.error_message = str(e)
            report_log.processing_time = (datetime.utcnow() - start_time).total_seconds()
        
        self.db.commit()
        return report_log
    
    async def _get_report_data(self, report_type: str, parameters: Dict[str, Any]) -> Union[List[Dict], Dict]:
        """Get data for report generation"""
        if report_type == ReportType.STUDENT_LIST:
            return await self._get_student_data(parameters)
        elif report_type == ReportType.TEACHER_LIST:
            return await self._get_teacher_data(parameters)
        elif report_type == ReportType.STATISTICS:
            return await self._get_statistics_data(parameters)
        else:
            raise ValueError(f"Unsupported report type: {report_type}")
    
    async def _get_student_data(self, parameters: Dict[str, Any]) -> List[Dict]:
        """Get student data"""
        query = self.db.query(Student).join(User)
        
        if parameters.get("class_id"):
            query = query.filter(Student.class_id == parameters["class_id"])
        
        students = query.all()
        
        return [
            {
                "id": student.id,
                "name": f"{student.user.first_name} {student.user.last_name}",
                "email": student.user.email,
                "class": student.class_.name if student.class_ else "Unassigned",
                "status": student.status
            }
            for student in students
        ]
    
    async def _get_teacher_data(self, parameters: Dict[str, Any]) -> List[Dict]:
        """Get teacher data"""
        query = self.db.query(Teacher).join(User)
        teachers = query.all()
        
        return [
            {
                "id": teacher.id,
                "name": f"{teacher.user.first_name} {teacher.user.last_name}",
                "email": teacher.user.email,
                "subject": teacher.subject,
                "status": teacher.status
            }
            for teacher in teachers
        ]
    
    async def _get_statistics_data(self, parameters: Dict[str, Any]) -> Dict[str, Any]:
        """Get statistics data"""
        total_students = self.db.query(Student).count()
        total_teachers = self.db.query(Teacher).count()
        total_classes = self.db.query(Class).count()
        
        return {
            "students": {"total": total_students},
            "teachers": {"total": total_teachers},
            "classes": {"total": total_classes}
        }
    
    async def _generate_output(self, data: Union[List[Dict], Dict], output_format: str, report_name: str) -> tuple[Path, int]:
        """Generate output file"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_name}_{timestamp}"
        
        if output_format == ReportFormat.PDF:
            return await self._generate_pdf(data, filename)
        elif output_format == ReportFormat.EXCEL:
            return await self._generate_excel(data, filename)
        elif output_format == ReportFormat.CSV:
            return await self._generate_csv(data, filename)
        elif output_format == ReportFormat.JSON:
            return await self._generate_json(data, filename)
        else:
            raise ValueError(f"Unsupported format: {output_format}")
    
    async def _generate_pdf(self, data: Union[List[Dict], Dict], filename: str) -> tuple[Path, int]:
        """Generate PDF report"""
        file_path = self.report_dir / f"{filename}.pdf"
        
        doc = SimpleDocTemplate(str(file_path), pagesize=A4)
        story = []
        
        title = Paragraph("Report", self.styles['Heading1'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        if isinstance(data, list) and data:
            headers = list(data[0].keys())
            table_data = [headers]
            
            for row in data:
                table_data.append([str(row.get(header, "")) for header in headers])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        
        doc.build(story)
        file_size = file_path.stat().st_size
        return file_path, file_size
    
    async def _generate_excel(self, data: Union[List[Dict], Dict], filename: str) -> tuple[Path, int]:
        """Generate Excel report"""
        file_path = self.report_dir / f"{filename}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if isinstance(data, list) and data:
            df = pd.DataFrame(data)
            
            for col_num, column in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num, value=column)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF")
            
            for row_num, row in enumerate(df.values, 2):
                for col_num, value in enumerate(row, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
        
        wb.save(str(file_path))
        file_size = file_path.stat().st_size
        return file_path, file_size
    
    async def _generate_csv(self, data: Union[List[Dict], Dict], filename: str) -> tuple[Path, int]:
        """Generate CSV report"""
        file_path = self.report_dir / f"{filename}.csv"
        
        if isinstance(data, list):
            df = pd.DataFrame(data)
        else:
            flat_data = [[key, value] for key, value in data.items()]
            df = pd.DataFrame(flat_data, columns=["Metric", "Value"])
        
        df.to_csv(file_path, index=False)
        file_size = file_path.stat().st_size
        return file_path, file_size
    
    async def _generate_json(self, data: Union[List[Dict], Dict], filename: str) -> tuple[Path, int]:
        """Generate JSON report"""
        file_path = self.report_dir / f"{filename}.json"
        
        report_data = {
            "generated_at": datetime.now().isoformat(),
            "data": data
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, default=str)
        
        file_size = file_path.stat().st_size
        return file_path, file_size
    
    async def get_report_logs(self, skip: int = 0, limit: int = 100) -> List[ReportLog]:
        """Get report logs"""
        return self.db.query(ReportLog).order_by(
            ReportLog.created_at.desc()
        ).offset(skip).limit(limit).all()
    
    async def get_report_statistics(self) -> Dict[str, Any]:
        """Get report statistics"""
        total_reports = self.db.query(ReportLog).count()
        completed_reports = self.db.query(ReportLog).filter(
            ReportLog.status == ReportStatus.COMPLETED
        ).count()
        
        return {
            "total_reports": total_reports,
            "completed_reports": completed_reports,
            "success_rate": round((completed_reports / total_reports * 100) if total_reports > 0 else 0, 2)
        } 